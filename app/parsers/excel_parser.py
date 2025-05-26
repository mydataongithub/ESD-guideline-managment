# app/parsers/excel_parser.py
import io
import openpyxl
from typing import Dict, Any, List, Optional
import pandas as pd
from .base_parser import BaseParser


class ExcelParser(BaseParser):
    """Parser to extract rule information from Excel documents."""
    
    def __init__(self, file_path: Optional[str] = None, file_content: Optional[bytes] = None):
        super().__init__(file_path, file_content)
        # Load workbook from file_path or file_content
        if file_path:
            self.workbook = openpyxl.load_workbook(file_path, data_only=True)
        else:
            self.workbook = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        
        # Also load with pandas for easier data processing
        if file_path:
            self.dataframes = pd.read_excel(file_path, sheet_name=None)
        else:
            self.dataframes = pd.read_excel(io.BytesIO(file_content), sheet_name=None)
            
    def extract_rules(self) -> List[Dict[str, Any]]:
        """
        Extract rules from Excel sheets.
        Looks for sheets with naming patterns like "Rules", "ESD Rules", "Latchup Rules".
        Also tries to automatically detect tables with rule-like contents.
        
        Returns:
            List of dictionaries with rule data
        """
        rules = []
        
        # Try to identify rule tables in all sheets
        for sheet_name, df in self.dataframes.items():
            # Process sheets with names that suggest they contain rules
            if any(rule_keyword in sheet_name.lower() for rule_keyword in ["rule", "esd", "latchup", "guideline"]):
                # Look for columns that might contain rule information
                df.columns = [str(col).lower() for col in df.columns]  # Normalize column names
                
                # Try to map columns to rule properties
                column_mapping = self._detect_column_mapping(df)
                
                # Extract rules using the mapping
                if column_mapping:
                    for _, row in df.iterrows():
                        rule = self._extract_rule_from_row(row, column_mapping)
                        if rule:
                            rules.append(rule)
        
        return rules
    
    def _detect_column_mapping(self, df) -> Dict[str, str]:
        """
        Try to map DataFrame columns to rule properties.
        
        Args:
            df: pandas DataFrame
        
        Returns:
            Dictionary mapping rule properties to column names
        """
        mapping = {}
        columns = [col.lower() for col in df.columns]
        
        # Map common column patterns to rule properties
        possible_mappings = {
            "title": ["title", "name", "rule name", "rule", "rule title", "rule id", "id"],
            "content": ["content", "description", "rule", "text", "rule content", "rule text", "requirement"],
            "rule_type": ["type", "rule type", "category", "classification"],
            "severity": ["severity", "priority", "importance", "criticality"],
            "category": ["category", "group", "area", "domain"],
        }
        
        for prop, possible_cols in possible_mappings.items():
            for col in columns:
                if any(possible == col or possible in col for possible in possible_cols):
                    mapping[prop] = col
                    break
        
        return mapping
    
    def _extract_rule_from_row(self, row, column_mapping) -> Optional[Dict[str, Any]]:
        """
        Extract a rule from a DataFrame row using the column mapping.
        
        Args:
            row: pandas Series (DataFrame row)
            column_mapping: Dictionary mapping rule properties to column names
            
        Returns:
            Dictionary with rule data or None if required fields missing
        """
        rule = {}
        
        # Get required fields
        for prop in ["title", "content"]:
            if prop in column_mapping:
                rule[prop] = str(row[column_mapping[prop]])
            else:
                # If we can't find required fields, can't create a valid rule
                return None
        
        # Get optional fields
        for prop in ["rule_type", "severity", "category"]:
            if prop in column_mapping:
                rule[prop] = str(row[column_mapping[prop]])
        
        # Map rule_type to enum values if applicable
        if "rule_type" in rule:
            rule_type = rule["rule_type"].lower()
            if "esd" in rule_type:
                rule["rule_type"] = "esd"
            elif "latchup" in rule_type:
                rule["rule_type"] = "latchup"
            else:
                rule["rule_type"] = "general"
                
        return rule
    
    def extract_metadata(self) -> Dict[str, Any]:
        """
        Extract metadata from the Excel document.
        
        Returns:
            Dictionary with metadata
        """
        metadata = {}
        
        # Get document properties
        props = self.workbook.properties
        
        # Extract available metadata
        if props.title:
            metadata["title"] = props.title
        if props.subject:
            metadata["subject"] = props.subject
        if props.creator:
            metadata["author"] = props.creator
        if props.created:
            metadata["created"] = props.created
        if props.modified:
            metadata["modified"] = props.modified
            
        # Get sheet names
        metadata["sheets"] = self.workbook.sheetnames
        
        return metadata
    
    def extract_images(self) -> List[Dict[str, Any]]:
        """
        Extract images from the Excel document.
        
        Returns:
            List of dictionaries with image data
        """
        images = []
        
        # Excel images are more complex to extract,
        # need to iterate through drawing shapes in each sheet
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            
            if sheet._images:
                for i, img in enumerate(sheet._images):
                    try:
                        image_data = {
                            "sheet": sheet_name,
                            "filename": f"image_{sheet_name}_{i}.png",
                            "image_data": img._data(),
                            "mime_type": "image/png"  # Excel typically stores images as PNG
                        }
                        images.append(image_data)
                    except Exception:
                        # Skip images that can't be processed
                        pass
        
        return images
